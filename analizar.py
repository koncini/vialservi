#!/usr/bin/python

from tkinter import *
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.comments import Comment
import os.path

root = Tk()
threshold_a = Spinbox(root, from_=0, to=100000)
threshold_b = Spinbox(root, from_=0, to=100000)
current_path = ""


def start_gui():
    root.title('Analizar reporte Vialservi')
    root.iconbitmap('favicon.ico')
    root.geometry("300x80")
    root.resizable(0, 0)
    high_threshold_label = Label(root,  text="Límite máximo de desajuste")
    mid_threshold_label = Label(root, text="Límite medio de desajuste")
    analyze = Button(root, text="Analizar Reporte", command=get_file)
    high_threshold_label.grid(row=0, sticky=W)
    mid_threshold_label.grid(row=1, sticky=W)
    threshold_a.grid(row=0, column=1, sticky=E)
    threshold_b.grid(row=1, column=1, sticky=E)
    analyze.grid(columnspan=2)


def get_file():
    global current_path
    root.withdraw()
    root.update()
    path_string = askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    current_path = os.path.dirname(path_string)
    if path_string != "":
        work_book = load_workbook(path_string)
        analyze_file(work_book)


def get_sheet(wb, index):
    sheet_names = wb.sheetnames
    sheet = wb[sheet_names[index]]
    return sheet


def get_sr_record(sheet, record_col, value_col):
    values = []
    fields = []
    current_record = {}
    last_recorded_id_value = None
    last_recorded_value = sheet["O2"].value
    last_recorded_id_coordinate = "F2"
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=value_col):
        values.append(last_recorded_value)
        fields.append(last_recorded_id_coordinate)
        if last_recorded_id_value != row[record_col-1].value:
            current_total_value = sum(values)
            current_record.update({last_recorded_id_value: (fields, current_total_value)})
            values = []
            fields = []
        last_recorded_id_value = row[record_col-1].value
        last_recorded_id_coordinate = row[record_col-1].coordinate
        last_recorded_value = row[value_col-1].value
    return current_record


def get_vs_record(sheet, record_col, value_col):
    current_record = {}
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=value_col):
        current_record.update({row[record_col-1].value: (row[record_col-1].coordinate, row[value_col-1].value)})
    return current_record


def paint_row(ws, cell_value, color_value):
    fill = PatternFill(start_color=color_value, end_color=color_value, fill_type="solid")
    str_coordinate = cell_value.coordinate
    xy = coordinate_from_string(str_coordinate)
    row_index = xy[1]
    for rows in ws.iter_rows(min_row=row_index, max_row=row_index, min_col=1, max_col=ws.max_column):
        for cell in rows:
            cell.fill = fill


def analyze_file(file):
    high_threshold = int(threshold_a.get())
    mid_threshold = int(threshold_b.get())
    current_wb = file
    ag_sheet = get_sheet(current_wb, 0)
    cs_sheet = get_sheet(current_wb, 1)
    ag_record = get_vs_record(ag_sheet, 3, 10)
    cs_record = get_sr_record(cs_sheet, 6, 15)
    for key in ag_record:
        if key in cs_record:
            cs_match_data = cs_record[key]
            ag_match_data = ag_record[key]
            match_diff = abs(ag_match_data[1]-cs_match_data[1])
            if match_diff >= high_threshold:
                warning_record = ag_sheet[ag_match_data[0]]
                comment = Comment(u'Presenta desajuste por: $'+ str(match_diff), u'Análisis')
                warning_record.comment = comment
                paint_row(ag_sheet, warning_record, "FF0000")
                for warning_match in cs_match_data[0]:
                    warning_record = cs_sheet[warning_match]
                    paint_row(cs_sheet, warning_record, "FF0000")
            elif mid_threshold <= match_diff <= high_threshold:
                warning_record = ag_sheet[ag_match_data[0]]
                comment = Comment(u'Presenta desajuste por: $' + str(match_diff), u'Análisis')
                warning_record.comment = comment
                paint_row(ag_sheet, warning_record, "FFFF00")
                for warning_match in cs_match_data[0]:
                    warning_record = cs_sheet[warning_match]
                    paint_row(cs_sheet, warning_record, "FFFF00")
            else:
                warning_record = ag_sheet[ag_match_data[0]]
                paint_row(ag_sheet, warning_record, "00FF00")
                for warning_match in cs_match_data[0]:
                    warning_record = cs_sheet[warning_match]
                    paint_row(cs_sheet, warning_record, "00FF00")
    destination = os.path.join(current_path, 'Reporte Analizado.xlsx')
    current_wb.save(destination)


if __name__ == "__main__":
    start_gui()
    root.mainloop()
